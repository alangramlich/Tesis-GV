# -*- coding: utf-8 -*-
"""
Created on Sat Jul 12 14:59:38 2025

@author: alang
"""

# -*- coding: utf-8 -*-
"""
Elimina ventanas bufferzone por dataset y genera control visual por hoja
"""

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------- CONFIGURACIÓN ----------
# 👇 MODIFICÁ esta lista para ajustar los parámetros de cada archivo
configuraciones = [
    {"archivo": "dataset1.xlsx", "ventanas_a_eliminar_antes": 2, "ventanas_a_eliminar_despues": 3},
    {"archivo": "dataset2.xlsx", "ventanas_a_eliminar_antes": 2, "ventanas_a_eliminar_despues": 5},
    {"archivo": "dataset3.xlsx", "ventanas_a_eliminar_antes": 2, "ventanas_a_eliminar_despues": 2},
    {"archivo": "dataset4.xlsx", "ventanas_a_eliminar_antes": 2, "ventanas_a_eliminar_despues": 3},
    {"archivo": "dataset5.xlsx", "ventanas_a_eliminar_antes": 2, "ventanas_a_eliminar_despues": 3},
    {"archivo": "dataset6.xlsx", "ventanas_a_eliminar_antes": 2, "ventanas_a_eliminar_despues": 3}
]
# ----------------------------------

def marcar_exclusiones(df, ventanas_a_eliminar_antes, ventanas_a_eliminar_despues):
    indices_a_excluir = set()
    etiquetas = df["clasificacion"].values

    for i in range(len(etiquetas)):
        if etiquetas[i] == "preictal":
            if i == 0 or etiquetas[i - 1] != "preictal":
                for j in range(1, ventanas_a_eliminar_antes + 1):
                    if i - j >= 0 and etiquetas[i - j] == "no_preictal":
                        indices_a_excluir.add(i - j)
            if i == len(etiquetas) - 1 or etiquetas[i + 1] != "preictal":
                for j in range(1, ventanas_a_eliminar_despues + 1):
                    if i + j < len(etiquetas) and etiquetas[i + j] == "no_preictal":
                        indices_a_excluir.add(i + j)

    resumen = {
        "cantidad_eliminadas": len(indices_a_excluir),
        "primer_indice": min(indices_a_excluir) if indices_a_excluir else None,
        "ultimo_indice": max(indices_a_excluir) if indices_a_excluir else None,
        "indices_eliminados": sorted(indices_a_excluir)
    }

    df_filtrado = df.drop(index=sorted(indices_a_excluir)).reset_index(drop=True)
    return df_filtrado, resumen

for cfg in configuraciones:
    archivo = cfg["archivo"]
    ventanas_a_eliminar_antes = cfg["ventanas_a_eliminar_antes"]
    ventanas_a_eliminar_despues = cfg["ventanas_a_eliminar_despues"]

    print(f"\n📂 Procesando {archivo}...")

    hojas = pd.ExcelFile(archivo).sheet_names
    hojas = [h for h in hojas if h != "Hoja1"]

    resultado = {}
    resumen_por_hoja = {}
    control_dict = {}

    for hoja in hojas:
        df = pd.read_excel(archivo, sheet_name=hoja)
        df_filtrado, resumen = marcar_exclusiones(df, ventanas_a_eliminar_antes, ventanas_a_eliminar_despues)

        resultado[hoja] = df_filtrado
        resumen_por_hoja[hoja] = resumen

        # Crear DataFrame para control visual
        df_control = df.copy()
        df_control["marcado"] = ["eliminado" if i in resumen["indices_eliminados"] else "ok" for i in range(len(df))]
        control_dict[hoja] = df_control

    # Guardar archivo filtrado
    nombre_salida = archivo.replace(".xlsx", "_filtrado_buffer.xlsx")
    with pd.ExcelWriter(nombre_salida) as writer:
        for hoja, data in resultado.items():
            data.to_excel(writer, sheet_name=hoja, index=False)

    print(f"✅ Guardado filtrado: {nombre_salida}")

    # Guardar archivo de control visual
    nombre_control = archivo.replace(".xlsx", "_control.xlsx")
    with pd.ExcelWriter(nombre_control, engine="openpyxl") as writer:
        for hoja, data in control_dict.items():
            data.to_excel(writer, sheet_name=hoja, index=False)

    # Pintar en rojo los eliminados
    wb = load_workbook(nombre_control)
    rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for hoja in wb.sheetnames:
        ws = wb[hoja]
        col_marcado = None
        for i, cell in enumerate(ws[1]):
            if cell.value == "marcado":
                col_marcado = i + 1
                break
        if col_marcado:
            for row in ws.iter_rows(min_row=2, min_col=col_marcado, max_col=col_marcado):
                for cell in row:
                    if cell.value == "eliminado":
                        for c in ws[cell.row]:
                            c.fill = rojo

    wb.save(nombre_control)
    print(f"📝 Guardado control: {nombre_control}")

    # Imprimir resumen
    print("📊 Resumen por hoja:")
    for hoja, resumen in resumen_por_hoja.items():
        print(f"  • {hoja}:")
        print(f"     - Ventanas eliminadas: {resumen['cantidad_eliminadas']}")
        print(f"     - Primer índice eliminado: {resumen['primer_indice']}")
        print(f"     - Último índice eliminado: {resumen['ultimo_indice']}")
